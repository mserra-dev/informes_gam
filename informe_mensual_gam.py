#!/usr/bin/env python3
"""
Informes mensuales automáticos — El Litoral · Google Ad Manager
================================================================
Genera y envía por email dos informes el 1ro de cada mes:
  1. Pautas Venta Directa   (anunciante + orden, ingresos ARS)
  2. CTR y Viewability      (bloques / ad units, CTR + ActiveView)

Uso:
  python informe_mensual_gam.py              # ambos informes
  python informe_mensual_gam.py --pautas     # solo pautas
  python informe_mensual_gam.py --bloques    # solo bloques

Cron (1ro de cada mes a las 8:00 AM):
  0 8 1 * * /usr/bin/python3 /ruta/informe_mensual_gam.py >> /var/log/gam_informes.log 2>&1
"""

import argparse
import base64
import csv
import datetime
import io
import os
import time
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

import openpyxl
from googleads import ad_manager
from google.oauth2 import service_account
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ──────────────────────────────────────────────────────────────
# CONFIGURACIÓN  (ajustar antes de correr)
# ──────────────────────────────────────────────────────────────
GAM_NETWORK_CODE  = "21877992475"
# Local: archivo service_account.json junto al script
# GitHub Actions: variable de entorno GAM_SA_PATH apunta a /tmp/gam_sa.json
SERVICE_ACCOUNT_JSON = os.getenv(
    "GAM_SA_PATH",
    os.path.join(os.path.dirname(__file__), "service_account.json")
)

DRIVE_FOLDER_ID   = "1UsYfdcsKCiVJJ_VVBvaNkCHL9bXeSOY-"
EMAIL_RECIPIENT   = "matias.serra@arcadiaconsultora.com"
# Si el service account tiene delegación de dominio, descomentá la siguiente línea:
# IMPERSONATE_USER = "matias.serra@arcadiaconsultora.com"
IMPERSONATE_USER  = None

GAM_API_VERSION   = "v202411"

# ──────────────────────────────────────────────────────────────
# ESTILOS EXCEL (compartidos)
# ──────────────────────────────────────────────────────────────
C_TITLE_BG  = "0D2244"
C_HEADER_BG = "1B3A6B"
C_SUB_BG    = "EEF1F7"
C_ALT       = "F4F7FB"
C_BORDER    = "B0B8C8"


def thin():
    s = Side(style="thin", color=C_BORDER)
    return Border(left=s, right=s, top=s, bottom=s)


# ──────────────────────────────────────────────────────────────
# HELPERS GENERALES
# ──────────────────────────────────────────────────────────────
def last_month_info():
    """Devuelve (start, end, nombre_mes_es, año) del mes anterior."""
    today = datetime.date.today()
    end   = today.replace(day=1) - datetime.timedelta(days=1)
    start = end.replace(day=1)
    mes   = {
        1: "Enero", 2: "Febrero", 3: "Marzo", 4: "Abril",
        5: "Mayo",  6: "Junio",   7: "Julio", 8: "Agosto",
        9: "Septiembre", 10: "Octubre", 11: "Noviembre", 12: "Diciembre",
    }[end.month]
    return start, end, mes, end.year


def gam_client() -> ad_manager.AdManagerClient:
    """Crea el cliente de GAM usando el service account."""
    return ad_manager.AdManagerClient.LoadFromStorage(SERVICE_ACCOUNT_JSON)


def google_creds(scopes: list[str]):
    """Credenciales de Google API para Drive y Gmail."""
    creds = service_account.Credentials.from_service_account_file(
        SERVICE_ACCOUNT_JSON, scopes=scopes
    )
    if IMPERSONATE_USER:
        creds = creds.with_subject(IMPERSONATE_USER)
    return creds


def wait_for_report(report_service, job_id: int) -> None:
    """Polling hasta que el reporte de GAM esté listo."""
    while True:
        status = report_service.getReportJobStatus(job_id)
        if status == "COMPLETED":
            return
        if status == "FAILED":
            raise RuntimeError(f"GAM report job {job_id} falló.")
        print(f"   · Esperando reporte GAM (status: {status})...")
        time.sleep(5)


def gam_date(d: datetime.date) -> dict:
    return {"year": d.year, "month": d.month, "day": d.day}


def upload_drive(file_path: str, file_name: str) -> str:
    """Sube (o actualiza) el archivo en Drive. Devuelve el file_id."""
    creds   = google_creds(["https://www.googleapis.com/auth/drive"])
    service = build("drive", "v3", credentials=creds)
    mime    = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    q        = f"name='{file_name}' and '{DRIVE_FOLDER_ID}' in parents and trashed=false"
    existing = service.files().list(q=q, fields="files(id)").execute().get("files", [])
    media    = MediaFileUpload(file_path, mimetype=mime)

    if existing:
        file_id = existing[0]["id"]
        service.files().update(fileId=file_id, media_body=media).execute()
        print(f"   · Drive actualizado: {file_id}")
    else:
        meta    = {"name": file_name, "parents": [DRIVE_FOLDER_ID]}
        result  = service.files().create(body=meta, media_body=media, fields="id").execute()
        file_id = result["id"]
        print(f"   · Drive nuevo: {file_id}")

    return file_id


def send_email(subject: str, html_body: str) -> None:
    """Envía email vía Gmail API."""
    creds   = google_creds(["https://www.googleapis.com/auth/gmail.send"])
    service = build("gmail", "v1", credentials=creds)
    msg     = MIMEMultipart("alternative")
    msg["Subject"] = subject
    msg["To"]      = EMAIL_RECIPIENT
    msg.attach(MIMEText(html_body, "html"))
    raw = base64.urlsafe_b64encode(msg.as_bytes()).decode()
    service.users().messages().send(userId="me", body={"raw": raw}).execute()
    print(f"   · Email enviado a {EMAIL_RECIPIENT}")


# ══════════════════════════════════════════════════════════════
# INFORME 1 — PAUTAS VENTA DIRECTA
# ══════════════════════════════════════════════════════════════
def run_pautas(start: datetime.date, end: datetime.date, mes: str, year: int) -> None:
    print("\n📋 Generando informe PAUTAS VENTA DIRECTA...")

    # 1. Reporte GAM
    client         = gam_client()
    report_service = client.GetService("ReportService", version=GAM_API_VERSION)

    job = report_service.runReportJob({
        "reportQuery": {
            "dimensions":    ["ADVERTISER_NAME", "ORDER_NAME"],
            "columns":       [
                "AD_SERVER_IMPRESSIONS",
                "AD_SERVER_CLICKS",
                "AD_SERVER_CTR",
                "AD_SERVER_REVENUE",
            ],
            "dateRangeType": "CUSTOM_DATE",
            "startDate":     gam_date(start),
            "endDate":       gam_date(end),
        }
    })
    wait_for_report(report_service, job["id"])

    import urllib.request
    url = report_service.getReportDownloadURL(job["id"], "CSV_DUMP")
    with urllib.request.urlopen(url) as r:
        csv_text = r.read().decode("utf-8")

    # 2. Parsear CSV
    reader = csv.DictReader(io.StringIO(csv_text))
    rows   = []
    for r in reader:
        advertiser = r.get("Dimension.ADVERTISER_NAME", "").strip()
        order      = r.get("Dimension.ORDER_NAME", "").strip()
        if advertiser.lower() in ("total", ""):
            continue
        impr    = int(float(r.get("Column.AD_SERVER_IMPRESSIONS", 0) or 0))
        clicks  = int(float(r.get("Column.AD_SERVER_CLICKS", 0) or 0))
        ctr_raw = r.get("Column.AD_SERVER_CTR", "0") or "0"
        ctr     = float(ctr_raw.replace("%", "")) / 100
        revenue = float(r.get("Column.AD_SERVER_REVENUE", 0) or 0)
        rows.append([advertiser, order, impr, clicks, ctr, revenue])

    rows.sort(key=lambda x: x[0])
    totals = [
        "TOTALES", "",
        sum(r[2] for r in rows),
        sum(r[3] for r in rows),
        sum(r[3] for r in rows) / max(sum(r[2] for r in rows), 1),
        sum(r[5] for r in rows),
    ]
    print(f"   · {len(rows)} filas | ARS {totals[5]:,.2f}")

    # 3. Excel
    file_name   = f"Informe_Pautas_VentaDirecta_{mes}{year}.xlsx"
    output_path = f"/tmp/{file_name}"
    _excel_pautas(rows, totals, start, end, mes, year, output_path)

    # 4. Drive
    file_id = upload_drive(output_path, file_name)

    # 5. Email
    html    = _email_pautas(mes, year, start, end, totals, rows, file_id)
    subject = f"📋 Informe Pautas Venta Directa - El Litoral - {mes} {year}"
    send_email(subject, html)


def _excel_pautas(rows, totals, start, end, mes, year, path):
    UMBRAL  = 1_000_000
    aligns  = ["left", "left", "right", "right", "right", "right"]
    formats = ["", "", "#,##0", "#,##0", "0.00%", '#,##0.00 "ARS"']

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Pautas {mes} {year}"

    for i, w in enumerate([38, 44, 18, 12, 12, 22], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value = f"INFORME DE PAUTAS — Venta Directa | El Litoral Ad Manager | {mes} {year}"
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:F2")
    c = ws["A2"]
    c.value = (f"Período: {start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')}  |  "
               "Fuente: GAM API  |  Ad Server (venta directa)  |  Moneda: ARS")
    c.font      = Font(name="Calibri", italic=True, size=9, color="5A6A8A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=C_SUB_BG)
    ws.row_dimensions[2].height = 18

    headers = ["Anunciante", "Pedido / Campaña", "Impresiones", "Clics", "CTR", "Ingresos (ARS)"]
    ws.row_dimensions[3].height = 36
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin()

    for i, row in enumerate(rows):
        r          = i + 4
        ws.row_dimensions[r].height = 18
        has_rev    = row[5] >= UMBRAL
        bg         = "FFFDE7" if has_rev else (C_ALT if i % 2 == 0 else "FFFFFF")
        for col, (val, fmt, aln) in enumerate(zip(row, formats, aligns), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Calibri", size=10,
                               bold=(has_rev and col == 6),
                               color=("1A5C1A" if (has_rev and col == 6) else "000000"))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=aln, vertical="center", indent=(1 if col <= 2 else 0))
            if fmt:
                c.number_format = fmt
            c.border = thin()

    r_tot = len(rows) + 4
    ws.row_dimensions[r_tot].height = 22
    for col, (val, fmt, aln) in enumerate(zip(totals, formats, aligns), 1):
        c = ws.cell(row=r_tot, column=col, value=val)
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal=aln, vertical="center", indent=(1 if col == 1 else 0))
        if fmt:
            c.number_format = fmt
        c.border = thin()

    ws.freeze_panes = "A4"

    note_row = r_tot + 2
    ws.merge_cells(f"A{note_row}:F{note_row}")
    note = ws.cell(row=note_row, column=1,
        value=(f"Nota: Filas en amarillo tienen ingresos ≥ ARS 1.000.000. "
               "Pautas con ARS 0,00 no tienen ingresos asignados en GAM. "
               f"Generado automáticamente el {datetime.date.today().strftime('%d/%m/%Y')}."))
    note.font      = Font(name="Calibri", italic=True, size=8, color="7A8AA0")
    note.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 28

    wb.save(path)


def _email_pautas(mes, year, start, end, totals, rows, file_id):
    top5 = sorted([r for r in rows if r[5] > 0], key=lambda x: -x[5])[:5]
    icons = ["🏆", "🥈", "🥉", "4️⃣", "5️⃣"]
    top5_html = ""
    for i, r in enumerate(top5):
        bg = "#FAFBFD" if i % 2 else "#FFFFFF"
        top5_html += (
            f'<tr style="background:{bg};border-bottom:1px solid #EEF1F7;">'
            f'<td style="padding:8px 10px;font-size:13px;color:#222;">{icons[i]} {r[0]}</td>'
            f'<td style="padding:8px 10px;font-size:13px;font-weight:700;color:#1A5C1A;text-align:right;">$ {r[5]:,.0f}</td>'
            f'<td style="padding:8px 10px;font-size:12px;color:#4A5568;text-align:right;">{r[2]:,}</td>'
            f"</tr>"
        )
    drive_link = f"https://drive.google.com/file/d/{file_id}/view"
    today_str  = datetime.date.today().strftime("%d/%m/%Y")
    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#F0F4F8;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#F0F4F8;padding:24px 0;">
<tr><td align="center"><table width="620" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1);">
<tr><td style="background:linear-gradient(135deg,#0D2244,#1B3A6B);padding:28px 32px;">
  <p style="margin:0 0 4px;font-size:11px;color:#8BAAD4;text-transform:uppercase;letter-spacing:1.5px;">El Litoral · Google Ad Manager</p>
  <h1 style="margin:0;font-size:22px;font-weight:700;color:#fff;">Informe Pautas Venta Directa</h1>
  <p style="margin:6px 0 0;font-size:14px;color:#A8C4E0;">Período: 1 al {end.day} de {mes.lower()} de {year}</p>
</td></tr>
<tr><td style="padding:24px 32px 8px;">
  <p style="margin:0;font-size:13.5px;color:#4A5568;line-height:1.7;">
    Resumen de <strong>pautas de venta directa</strong> (Ad Server) en GAM. <strong>{len(rows)} órdenes</strong> activas.
  </p>
</td></tr>
<tr><td style="padding:16px 32px;">
  <table width="100%" cellpadding="0" cellspacing="8">
    <tr>
      <td width="50%" style="padding-right:6px;"><div style="background:#EEF1F7;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #1B3A6B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">Impresiones</p>
        <p style="margin:0;font-size:26px;font-weight:700;color:#0D2244;">{totals[2]:,.0f}</p>
      </div></td>
      <td width="50%" style="padding-left:6px;"><div style="background:#EEF1F7;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #1B3A6B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">Clics</p>
        <p style="margin:0;font-size:26px;font-weight:700;color:#0D2244;">{totals[3]:,.0f}</p>
      </div></td>
    </tr>
    <tr>
      <td width="50%" style="padding-right:6px;padding-top:8px;"><div style="background:#EEF1F7;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #2A7A3B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">CTR promedio</p>
        <p style="margin:0;font-size:26px;font-weight:700;color:#1A5C1A;">{totals[4]*100:.2f}%</p>
      </div></td>
      <td width="50%" style="padding-left:6px;padding-top:8px;"><div style="background:#FFF9E6;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #D4900A;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">Ingresos ARS</p>
        <p style="margin:0;font-size:22px;font-weight:700;color:#8B5E00;">$ {totals[5]:,.0f}</p>
      </div></td>
    </tr>
  </table>
</td></tr>
<tr><td style="padding:8px 32px 16px;">
  <p style="margin:0 0 12px;font-size:13px;font-weight:700;color:#1B3A6B;text-transform:uppercase;letter-spacing:.8px;border-bottom:2px solid #EEF1F7;padding-bottom:8px;">Top anunciantes por ingresos</p>
  <table width="100%" cellpadding="0" cellspacing="0">
    <tr style="background:#EEF1F7;">
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;">Anunciante</td>
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;text-align:right;">Ingresos ARS</td>
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;text-align:right;">Impr.</td>
    </tr>
    {top5_html}
    <tr style="background:#FAFBFD;">
      <td style="padding:8px 10px;font-size:12px;color:#6B7A99;font-style:italic;" colspan="3">+ más órdenes · Ver detalle completo en el archivo adjunto</td>
    </tr>
  </table>
</td></tr>
<tr><td style="padding:0 32px 24px;text-align:center;">
  <a href="{drive_link}" style="display:inline-block;background:linear-gradient(135deg,#0D2244,#1B3A6B);color:#fff;text-decoration:none;font-size:14px;font-weight:600;padding:13px 32px;border-radius:6px;">📥 Ver informe completo en Google Drive</a>
</td></tr>
<tr><td style="background:#F7F9FC;border-top:1px solid #E2E8F0;padding:16px 32px;">
  <p style="margin:0;font-size:11px;color:#8A9AB8;line-height:1.6;">
    Fuente: GAM API · Ad Server (venta directa) · Red El Litoral (ID: 21877992475)<br>
    Moneda: ARS · Generado automáticamente el {today_str}
  </p>
</td></tr>
</table></td></tr></table>
</body></html>"""


# ══════════════════════════════════════════════════════════════
# INFORME 2 — CTR Y VIEWABILITY (BLOQUES / AD UNITS)
# ══════════════════════════════════════════════════════════════
def run_bloques(start: datetime.date, end: datetime.date, mes: str, year: int) -> None:
    print("\n📊 Generando informe CTR Y VIEWABILITY (bloques)...")

    client         = gam_client()
    report_service = client.GetService("ReportService", version=GAM_API_VERSION)

    job = report_service.runReportJob({
        "reportQuery": {
            "dimensions":    ["AD_UNIT_NAME"],
            "adUnitView":    "FLAT",
            "columns":       [
                "AD_SERVER_IMPRESSIONS",
                "AD_SERVER_CLICKS",
                "AD_SERVER_CTR",
                "TOTAL_ACTIVE_VIEW_MEASURABLE_IMPRESSIONS",
                "TOTAL_ACTIVE_VIEW_VIEWABLE_IMPRESSIONS",
                "TOTAL_ACTIVE_VIEW_VIEWABLE_IMPRESSIONS_RATE",
            ],
            "dateRangeType": "CUSTOM_DATE",
            "startDate":     gam_date(start),
            "endDate":       gam_date(end),
            # Filtra solo ad units hijos de El Litoral (ajustar según jerarquía real)
            "statement": {
                "query": "WHERE PARENT_AD_UNIT_ID = 21877992475"
            }
        }
    })
    wait_for_report(report_service, job["id"])

    import urllib.request
    url = report_service.getReportDownloadURL(job["id"], "CSV_DUMP")
    with urllib.request.urlopen(url) as r:
        csv_text = r.read().decode("utf-8")

    reader = csv.DictReader(io.StringIO(csv_text))
    rows   = []
    for r in reader:
        name = r.get("Dimension.AD_UNIT_NAME", "").strip()
        if name.lower() in ("total", ""):
            continue
        impr     = int(float(r.get("Column.AD_SERVER_IMPRESSIONS", 0) or 0))
        clicks   = int(float(r.get("Column.AD_SERVER_CLICKS", 0) or 0))
        ctr_raw  = r.get("Column.AD_SERVER_CTR", "0") or "0"
        ctr      = float(ctr_raw.replace("%", "")) / 100
        meas     = int(float(r.get("Column.TOTAL_ACTIVE_VIEW_MEASURABLE_IMPRESSIONS", 0) or 0))
        view     = int(float(r.get("Column.TOTAL_ACTIVE_VIEW_VIEWABLE_IMPRESSIONS", 0) or 0))
        view_raw = r.get("Column.TOTAL_ACTIVE_VIEW_VIEWABLE_IMPRESSIONS_RATE", "0") or "0"
        viewr    = float(view_raw.replace("%", "")) / 100
        rows.append([name, impr, clicks, ctr, meas, view, viewr])

    rows.sort(key=lambda x: -x[1])  # orden por impresiones desc
    print(f"   · {len(rows)} bloques")

    totals = [
        "TOTALES",
        sum(r[1] for r in rows),
        sum(r[2] for r in rows),
        sum(r[2] for r in rows) / max(sum(r[1] for r in rows), 1),
        sum(r[4] for r in rows),
        sum(r[5] for r in rows),
        sum(r[5] for r in rows) / max(sum(r[4] for r in rows), 1),
    ]

    file_name   = f"Informe_CTR_Viewability_{mes}{year}.xlsx"
    output_path = f"/tmp/{file_name}"
    _excel_bloques(rows, totals, start, end, mes, year, output_path)

    file_id = upload_drive(output_path, file_name)

    html    = _email_bloques(mes, year, start, end, totals, rows, file_id)
    subject = f"📊 Informe CTR y Viewability - Ads units - El Litoral - {mes} {year}"
    send_email(subject, html)


def _excel_bloques(rows, totals, start, end, mes, year, path):
    headers = [
        "Bloque / Ad Unit",
        "Impresiones",
        "Clics",
        "CTR",
        "Impr. Medibles (AV)",
        "Impr. Visibles (AV)",
        "Viewability %",
    ]
    aligns  = ["left", "right", "right", "right", "right", "right", "right"]
    formats = ["", "#,##0", "#,##0", "0.00%", "#,##0", "#,##0", "0.00%"]

    UMBRAL_VIEWABILITY = 0.70  # highlight si viewability >= 70%

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Bloques {mes} {year}"

    for i, w in enumerate([46, 18, 12, 10, 22, 22, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value = f"INFORME CTR Y VIEWABILITY — Bloques | El Litoral Ad Manager | {mes} {year}"
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:G2")
    c = ws["A2"]
    c.value = (f"Período: {start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')}  |  "
               "Fuente: GAM API  |  Ad Server  |  ActiveView Google")
    c.font      = Font(name="Calibri", italic=True, size=9, color="5A6A8A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=C_SUB_BG)
    ws.row_dimensions[2].height = 18

    ws.row_dimensions[3].height = 36
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin()

    for i, row in enumerate(rows):
        r          = i + 4
        ws.row_dimensions[r].height = 18
        high_view  = row[6] >= UMBRAL_VIEWABILITY
        bg         = "E8F5E9" if high_view else (C_ALT if i % 2 == 0 else "FFFFFF")
        for col, (val, fmt, aln) in enumerate(zip(row, formats, aligns), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Calibri", size=10,
                               bold=(high_view and col == 7),
                               color=("1A5C1A" if (high_view and col == 7) else "000000"))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=aln, vertical="center", indent=(1 if col == 1 else 0))
            if fmt:
                c.number_format = fmt
            c.border = thin()

    r_tot = len(rows) + 4
    ws.row_dimensions[r_tot].height = 22
    for col, (val, fmt, aln) in enumerate(zip(totals, formats, aligns), 1):
        c = ws.cell(row=r_tot, column=col, value=val)
        c.font      = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal=aln, vertical="center", indent=(1 if col == 1 else 0))
        if fmt:
            c.number_format = fmt
        c.border = thin()

    ws.freeze_panes = "A4"

    note_row = r_tot + 2
    ws.merge_cells(f"A{note_row}:G{note_row}")
    note = ws.cell(row=note_row, column=1,
        value=(f"Nota: Filas en verde tienen Viewability ≥ 70%. "
               "ActiveView mide solo impresiones donde el ad estuvo al menos 50% visible por 1 seg. "
               f"Generado automáticamente el {datetime.date.today().strftime('%d/%m/%Y')}."))
    note.font      = Font(name="Calibri", italic=True, size=8, color="7A8AA0")
    note.alignment = Alignment(horizontal="left", wrap_text=True)
    ws.row_dimensions[note_row].height = 28

    wb.save(path)


def _email_bloques(mes, year, start, end, totals, rows, file_id):
    top5 = sorted(rows, key=lambda x: -x[6])[:5]  # top viewability
    icons = ["🥇", "🥈", "🥉", "4️⃣", "5️⃣"]
    top5_html = ""
    for i, r in enumerate(top5):
        bg = "#FAFBFD" if i % 2 else "#FFFFFF"
        top5_html += (
            f'<tr style="background:{bg};border-bottom:1px solid #EEF1F7;">'
            f'<td style="padding:8px 10px;font-size:12px;color:#222;">{icons[i]} {r[0]}</td>'
            f'<td style="padding:8px 10px;font-size:13px;font-weight:700;color:#1A5C1A;text-align:right;">{r[6]*100:.1f}%</td>'
            f'<td style="padding:8px 10px;font-size:12px;color:#4A5568;text-align:right;">{r[3]*100:.2f}%</td>'
            f'<td style="padding:8px 10px;font-size:12px;color:#4A5568;text-align:right;">{r[1]:,}</td>'
            f"</tr>"
        )
    drive_link = f"https://drive.google.com/file/d/{file_id}/view"
    today_str  = datetime.date.today().strftime("%d/%m/%Y")
    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#F0F4F8;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#F0F4F8;padding:24px 0;">
<tr><td align="center"><table width="620" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1);">
<tr><td style="background:linear-gradient(135deg,#0D2244,#1B3A6B);padding:28px 32px;">
  <p style="margin:0 0 4px;font-size:11px;color:#8BAAD4;text-transform:uppercase;letter-spacing:1.5px;">El Litoral · Google Ad Manager</p>
  <h1 style="margin:0;font-size:22px;font-weight:700;color:#fff;">Informe CTR y Viewability</h1>
  <p style="margin:6px 0 0;font-size:14px;color:#A8C4E0;">Bloques / Ad Units · {mes} {year}</p>
</td></tr>
<tr><td style="padding:24px 32px 8px;">
  <p style="margin:0;font-size:13.5px;color:#4A5568;line-height:1.7;">
    Rendimiento de <strong>{len(rows)} bloques publicitarios</strong> del Ad Server durante {mes.lower()} {year}. Incluye CTR e indicadores de viewability (ActiveView).
  </p>
</td></tr>
<tr><td style="padding:16px 32px;">
  <table width="100%" cellpadding="0" cellspacing="8">
    <tr>
      <td width="50%" style="padding-right:6px;"><div style="background:#EEF1F7;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #1B3A6B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">Impresiones</p>
        <p style="margin:0;font-size:26px;font-weight:700;color:#0D2244;">{totals[1]:,.0f}</p>
      </div></td>
      <td width="50%" style="padding-left:6px;"><div style="background:#EEF1F7;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #2A7A3B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">CTR promedio</p>
        <p style="margin:0;font-size:26px;font-weight:700;color:#1A5C1A;">{totals[3]*100:.2f}%</p>
      </div></td>
    </tr>
    <tr>
      <td width="50%" style="padding-right:6px;padding-top:8px;"><div style="background:#EEF1F7;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #1B3A6B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">Impr. Medibles (AV)</p>
        <p style="margin:0;font-size:22px;font-weight:700;color:#0D2244;">{totals[4]:,.0f}</p>
      </div></td>
      <td width="50%" style="padding-left:6px;padding-top:8px;"><div style="background:#E8F5E9;border-radius:8px;padding:16px;text-align:center;border-left:4px solid #2A7A3B;">
        <p style="margin:0 0 4px;font-size:11px;color:#6B7A99;text-transform:uppercase;">Viewability promedio</p>
        <p style="margin:0;font-size:26px;font-weight:700;color:#1A5C1A;">{totals[6]*100:.1f}%</p>
      </div></td>
    </tr>
  </table>
</td></tr>
<tr><td style="padding:8px 32px 16px;">
  <p style="margin:0 0 12px;font-size:13px;font-weight:700;color:#1B3A6B;text-transform:uppercase;letter-spacing:.8px;border-bottom:2px solid #EEF1F7;padding-bottom:8px;">Top bloques por viewability</p>
  <table width="100%" cellpadding="0" cellspacing="0">
    <tr style="background:#EEF1F7;">
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;">Bloque</td>
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;text-align:right;">Viewability</td>
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;text-align:right;">CTR</td>
      <td style="padding:7px 10px;font-size:11px;font-weight:700;color:#4A5568;text-transform:uppercase;text-align:right;">Impr.</td>
    </tr>
    {top5_html}
  </table>
</td></tr>
<tr><td style="padding:0 32px 24px;text-align:center;">
  <a href="{drive_link}" style="display:inline-block;background:linear-gradient(135deg,#0D2244,#1B3A6B);color:#fff;text-decoration:none;font-size:14px;font-weight:600;padding:13px 32px;border-radius:6px;">📥 Ver informe completo en Google Drive</a>
</td></tr>
<tr><td style="background:#F7F9FC;border-top:1px solid #E2E8F0;padding:16px 32px;">
  <p style="margin:0;font-size:11px;color:#8A9AB8;line-height:1.6;">
    Fuente: GAM API · Ad Server · ActiveView Google · Red El Litoral (ID: 21877992475)<br>
    Generado automáticamente el {today_str}
  </p>
</td></tr>
</table></td></tr></table>
</body></html>"""


# ══════════════════════════════════════════════════════════════
# INFORME 3 — INGRESOS PROGRAMÁTICOS (AdSense + AdX)
# ══════════════════════════════════════════════════════════════
def run_programatica(start: datetime.date, end: datetime.date, mes: str, year: int) -> None:
    print("\n💹 Generando informe PROGRAMÁTICA (AdSense + AdX)...")

    client         = gam_client()
    report_service = client.GetService("ReportService", version=GAM_API_VERSION)

    job = report_service.runReportJob({
        "reportQuery": {
            "dimensions":    ["PROGRAMMATIC_CHANNEL"],
            "columns":       [
                "ADSENSE_LINE_ITEM_LEVEL_REVENUE",
                "AD_EXCHANGE_LINE_ITEM_LEVEL_REVENUE",
            ],
            "dateRangeType": "CUSTOM_DATE",
            "startDate":     gam_date(start),
            "endDate":       gam_date(end),
        }
    })
    wait_for_report(report_service, job["id"])

    import urllib.request
    url = report_service.getReportDownloadURL(job["id"], "CSV_DUMP")
    with urllib.request.urlopen(url) as r:
        csv_text = r.read().decode("utf-8")

    reader = csv.DictReader(io.StringIO(csv_text))
    rows   = []
    for r in reader:
        canal   = r.get("Dimension.PROGRAMMATIC_CHANNEL", "").strip()
        if canal.lower() in ("total", ""):
            continue
        adsense = float(r.get("Column.ADSENSE_LINE_ITEM_LEVEL_REVENUE", 0) or 0)
        adx     = float(r.get("Column.AD_EXCHANGE_LINE_ITEM_LEVEL_REVENUE", 0) or 0)
        rows.append([canal, adsense, adx])

    total_adsense = sum(r[1] for r in rows)
    total_adx     = sum(r[2] for r in rows)
    total_combined = total_adsense + total_adx
    print(f"   · AdSense: USD {total_adsense:.2f} | AdX: USD {total_adx:.2f} | Total: USD {total_combined:.2f}")

    file_name   = f"Informe_Programatica_{mes}{year}.xlsx"
    output_path = f"/tmp/{file_name}"
    _excel_programatica(rows, total_adsense, total_adx, total_combined, start, end, mes, year, output_path)

    file_id = upload_drive(output_path, file_name)

    html    = _email_programatica(mes, year, start, end, total_adsense, total_adx, total_combined, rows, file_id)
    subject = f"💹 Informe Ingresos Programáticos (AdSense + AdX) - El Litoral - {mes} {year}"
    send_email(subject, html)


def _excel_programatica(rows, total_adsense, total_adx, total_combined, start, end, mes, year, path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Programática {mes} {year}"

    col_widths = [36, 22, 20, 20, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Título
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value = f"INFORME INGRESOS PROGRAMÁTICOS — El Litoral Ad Manager | {mes} {year}"
    c.font      = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C_TITLE_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Subtítulo
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value = (f"Período: {start.strftime('%d/%m/%Y')} al {end.strftime('%d/%m/%Y')}  |  "
               "Fuente: GAM API  |  Moneda: USD")
    c.font      = Font(name="Calibri", italic=True, size=9, color="5A6A8A")
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill      = PatternFill("solid", fgColor=C_SUB_BG)
    ws.row_dimensions[2].height = 18

    # KPI resumen
    ws.merge_cells("A3:E3")
    c = ws["A3"]
    c.value = (f"  AdSense: USD {total_adsense:,.2f}     |     "
               f"Ad Exchange: USD {total_adx:,.2f}     |     "
               f"TOTAL: USD {total_combined:,.2f}")
    c.font      = Font(name="Calibri", bold=True, size=11, color=C_TITLE_BG)
    c.fill      = PatternFill("solid", fgColor="D6E4F0")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[3].height = 22

    # Headers
    headers = ["Canal GAM", "Plataforma", "Ingresos AdSense (USD)", "Ingresos AdX (USD)", "Total (USD)"]
    ws.row_dimensions[4].height = 36
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin()

    # Mapeo de canal GAM → nombre legible y plataforma
    CANAL_MAP = {
        "-":                "AdSense (backfill / tradicional)",
        "Open Auction":     "Ad Exchange – Subasta abierta",
        "Private Auction":  "Ad Exchange – Subasta privada",
        "Preferred Deal":   "Ad Exchange – Acuerdo preferente",
        "Programmatic Guaranteed": "Ad Exchange – Programática garantizada",
    }
    PLATFORM_MAP = {
        "-":                "AdSense",
        "Open Auction":     "AdX",
        "Private Auction":  "AdX",
        "Preferred Deal":   "AdX",
        "Programmatic Guaranteed": "AdX",
    }

    aligns  = ["left", "center", "right", "right", "right"]
    formats = ["", "", '#,##0.00 "USD"', '#,##0.00 "USD"', '#,##0.00 "USD"']
    BG_ADSENSE = "E8F5E9"
    BG_ADX     = "E3F2FD"

    for i, row in enumerate(rows):
        r        = i + 5
        canal    = row[0]
        adsense  = row[1]
        adx      = row[2]
        plat     = PLATFORM_MAP.get(canal, "AdX" if adx > 0 else "AdSense")
        label    = CANAL_MAP.get(canal, canal)
        total_r  = adsense + adx
        bg       = BG_ADSENSE if plat == "AdSense" else BG_ADX
        data_row = [label, plat, adsense, adx, total_r]

        ws.row_dimensions[r].height = 22
        for col, (val, fmt, aln) in enumerate(zip(data_row, formats, aligns), 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font      = Font(name="Calibri", size=11, bold=(col == 5 and total_r > 0))
            c.fill      = PatternFill("solid", fgColor=bg)
            c.alignment = Alignment(horizontal=aln, vertical="center", indent=(1 if col <= 2 else 0))
            if fmt:
                c.number_format = fmt
            c.border = thin()

    # Totales
    r_tot = len(rows) + 5
    ws.row_dimensions[r_tot].height = 26
    totals_row = ["TOTALES", "", total_adsense, total_adx, total_combined]
    for col, (val, fmt, aln) in enumerate(zip(totals_row, formats, aligns), 1):
        c = ws.cell(row=r_tot, column=col, value=val)
        c.font      = Font(name="Calibri", bold=True, size=12, color="FFFFFF")
        c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
        c.alignment = Alignment(horizontal=aln, vertical="center", indent=(1 if col <= 2 else 0))
        if fmt:
            c.number_format = fmt
        c.border = thin()

    # Participación
    ws.row_dimensions[r_tot + 1].height = 8
    r_sh = r_tot + 2
    ws.merge_cells(f"A{r_sh}:E{r_sh}")
    c = ws.cell(row=r_sh, column=1, value="Participación sobre total programático")
    c.font      = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
    c.fill      = PatternFill("solid", fgColor=C_HEADER_BG)
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[r_sh].height = 20

    share_rows = [
        ("AdSense",          total_adsense / max(total_combined, 0.01),  "E8F5E9"),
        ("Ad Exchange (AdX)",total_adx / max(total_combined, 0.01),      "E3F2FD"),
    ]
    for i, (label, pct, bg) in enumerate(share_rows):
        r = r_sh + 1 + i
        ws.row_dimensions[r].height = 20
        ws.cell(row=r, column=1, value=label).fill = PatternFill("solid", fgColor=bg)
        ws.cell(row=r, column=1).font = Font(name="Calibri", size=10)
        ws.cell(row=r, column=1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(row=r, column=1).border = thin()
        c = ws.cell(row=r, column=2, value=pct)
        c.font = Font(name="Calibri", bold=True, size=10)
        c.number_format = "0.00%"
        c.fill = PatternFill("solid", fgColor=bg)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin()
        ws.merge_cells(f"C{r}:E{r}")
        desc = ws.cell(row=r, column=3)
        desc.fill = PatternFill("solid", fgColor=bg)
        desc.border = thin()

    ws.freeze_panes = "A5"
    wb.save(path)


def _email_programatica(mes, year, start, end, total_adsense, total_adx, total_combined, rows, file_id):
    pct_adsense = total_adsense / max(total_combined, 0.01) * 100
    pct_adx     = total_adx    / max(total_combined, 0.01) * 100
    drive_link  = f"https://drive.google.com/file/d/{file_id}/view"
    today_str   = datetime.date.today().strftime("%d/%m/%Y")
    return f"""<!DOCTYPE html>
<html lang="es"><head><meta charset="UTF-8"></head>
<body style="margin:0;padding:0;background:#F0F4F8;font-family:'Segoe UI',Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#F0F4F8;padding:24px 0;">
<tr><td align="center"><table width="620" cellpadding="0" cellspacing="0"
  style="background:#fff;border-radius:8px;overflow:hidden;box-shadow:0 2px 12px rgba(0,0,0,.1);">
<tr><td style="background:linear-gradient(135deg,#0D2244,#1B3A6B);padding:28px 32px;">
  <p style="margin:0 0 4px;font-size:11px;color:#8BAAD4;text-transform:uppercase;letter-spacing:1.5px;">El Litoral · Google Ad Manager</p>
  <h1 style="margin:0;font-size:22px;font-weight:700;color:#fff;">Ingresos Programáticos</h1>
  <p style="margin:6px 0 0;font-size:14px;color:#A8C4E0;">AdSense + Ad Exchange (AdX) · {mes} {year}</p>
</td></tr>
<tr><td style="padding:16px 32px;">
  <table width="100%" cellpadding="0" cellspacing="8">
    <tr>
      <td width="50%" style="padding-right:6px;"><div style="background:#E8F5E9;border-radius:8px;padding:18px 16px;text-align:center;border-left:4px solid #2A7A3B;">
        <p style="margin:0 0 4px;font-size:11px;color:#4A7A57;text-transform:uppercase;font-weight:600;">AdSense</p>
        <p style="margin:0 0 2px;font-size:26px;font-weight:700;color:#1A5C1A;">USD {total_adsense:,.2f}</p>
        <p style="margin:0;font-size:11px;color:#6B9B7A;">Backfill tradicional · {pct_adsense:.1f}%</p>
      </div></td>
      <td width="50%" style="padding-left:6px;"><div style="background:#E3F2FD;border-radius:8px;padding:18px 16px;text-align:center;border-left:4px solid #1565C0;">
        <p style="margin:0 0 4px;font-size:11px;color:#1A4A8A;text-transform:uppercase;font-weight:600;">Ad Exchange (AdX)</p>
        <p style="margin:0 0 2px;font-size:26px;font-weight:700;color:#0D3A7A;">USD {total_adx:,.2f}</p>
        <p style="margin:0;font-size:11px;color:#4A6A9A;">Subasta abierta · {pct_adx:.1f}%</p>
      </div></td>
    </tr>
    <tr><td colspan="2" style="padding-top:10px;">
      <div style="background:#FFF9E6;border-radius:8px;padding:16px;text-align:center;border:2px solid #D4900A;">
        <p style="margin:0 0 4px;font-size:11px;color:#8B5E00;text-transform:uppercase;font-weight:600;">Total programático combinado</p>
        <p style="margin:0;font-size:32px;font-weight:700;color:#6B4400;">USD {total_combined:,.2f}</p>
      </div>
    </td></tr>
  </table>
</td></tr>
<tr><td style="padding:0 32px 24px;text-align:center;">
  <a href="{drive_link}" style="display:inline-block;background:linear-gradient(135deg,#0D2244,#1B3A6B);color:#fff;text-decoration:none;font-size:14px;font-weight:600;padding:13px 32px;border-radius:6px;">📥 Ver informe completo en Google Drive</a>
</td></tr>
<tr><td style="background:#F7F9FC;border-top:1px solid #E2E8F0;padding:16px 32px;">
  <p style="margin:0;font-size:11px;color:#8A9AB8;line-height:1.6;">
    Fuente: GAM API · Programática · Red El Litoral (ID: 21877992475)<br>
    Moneda: USD · Generado automáticamente el {today_str}
  </p>
</td></tr>
</table></td></tr></table>
</body></html>"""


# ──────────────────────────────────────────────────────────────
# ENTRY POINT
# ──────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(description="Informes mensuales GAM El Litoral")
    parser.add_argument("--pautas",       action="store_true", help="Solo informe de pautas")
    parser.add_argument("--bloques",      action="store_true", help="Solo informe de bloques")
    parser.add_argument("--programatica", action="store_true", help="Solo informe programática")
    args = parser.parse_args()

    start, end, mes, year = last_month_info()
    print(f"🗓  Período: {start} → {end}  ({mes} {year})")

    run_all = not args.pautas and not args.bloques and not args.programatica

    if args.pautas or run_all:
        run_pautas(start, end, mes, year)

    if args.bloques or run_all:
        run_bloques(start, end, mes, year)

    if args.programatica or run_all:
        run_programatica(start, end, mes, year)

    print("\n✅ Listo.")


if __name__ == "__main__":
    main()
